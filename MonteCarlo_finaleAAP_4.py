import random as rd
import math
import numpy as np
import matplotlib.pyplot as plt
import openpyxl as xl
from openpyxl import Workbook

discount_rate=0.1045
repet=10000
price_matrix=[]
mature_matrix=[]
election_y=3 #because of election at the end of the year
taxrate=0.21
for l in range(repet):
    mature_prob=0.0001
    mature=False
    FCFE=[]
    sgaY=3611
    cogsY=5468
    salesY=9743
    net_inc=[]
    netinc=0
    for i in range(100):
        #creating scenario of election to see impact on corporate taxes
        election_y=election_y+1
        prob_election=0.5
        if election_y==4:
            democrat_republican=np.random.uniform()
            election_y==0
            if democrat_republican>prob_election:
                taxrate=0.21
            else:
                taxrate=0.35
            
       #creating a probability of mature stage
        if mature==False:       
            mature_prob=mature_prob+0.001
            mature_rand=np.random.uniform()
            if mature_rand<mature_prob:
                mature=True
                mature_matrix.append(i)

        if mature==True:
            #sales parameters
            a1=0
            b1=0.015
            #cogs parameters
            c1=0.005
            d1=0.015
            #SG&A parameters
            e1=0.005
            f1=0.02
        else:
            #sales parameters
            a1=0.012
            b1=0.03
            #cogs parameters
            c1=-0.0002
            d1=0.015
            #SG&A parameters
            e1=-0.00025
            f1=0.02

        prob_swan=0.01
        a=np.random.uniform()
        if a<prob_swan:
            salesY=salesY*(1+rd.gauss(-0.15,0.01)) #blackswan scenario
        else:
            salesY=salesY*(1+rd.gauss(a1,b1))

        cogsY=salesY*0.55*(1+rd.gauss(c1,d1)) #COGS always 54%-57% of sales
        grossprof=salesY-cogsY
        sgaY=sgaY*(1+rd.gauss(e1,f1))
        opeinc=grossprof-sgaY

        taxamount=opeinc*taxrate 

        netinc=opeinc-taxamount 
        net_inc.append(netinc)   


    workcap=50      
    capex=75

    year=2020
    debt_time=3
    for j in net_inc:
        year=year+1
              
        CFO_bforWC=j+300    #  this is the Non-cash item in the CF statement that we assume constant (historically constant)
        # interests payments historically constant
        workcap=workcap*(1+rd.gauss(0.001,0.1)) #high stdev as we can see historically
        CFO_net=CFO_bforWC+workcap
        posprec=j
        capex=capex*(1+rd.gauss(0.001,0.1)) #high stdev as we can see historically

        if debt_time==0:
            debt=-100
        if debt_time==1:
            debt=-250
        if debt_time>1:
            debt=-50
        if debt_time>9:
            debt_time==0         
 
        fcfe=CFO_net-capex+debt
        T=year-2020
        pv=fcfe/((1+discount_rate)**T)
        FCFE.append(pv)
    
    price=np.sum(FCFE)/69.6
    if price>0:
        price_matrix.append(price)
    
wb= Workbook()
ws1=wb.create_sheet("Results")
raw=2
for display in price_matrix:
    ws1.cell(raw,1).value = display
    raw=raw+1
wb.save(filename="MonteCarlo_FCFE_AAP.xlsx")
print('final year is:', year)
target_price=np.mean(price_matrix) 
print('our target price is: ',target_price)
plt.grid()
plt.hist(price_matrix, bins=50, density=True)
plt.show()
print('the average of mature cycle is reached in year: ' , np.mean(mature_matrix))
#print(price_matrix)
